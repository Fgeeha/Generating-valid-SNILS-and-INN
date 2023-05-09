import sys
import os
from PyQt6.QtWidgets import QApplication, QWidget, QLabel, \
    QLineEdit, QPushButton, QMessageBox, \
    QProgressBar
from openpyxl import Workbook
from get_inn_snils import get_snils, get_random_inn_fl


class ExcelForm(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Формирование excel файлов')
        self.resize(350, 200)

        self.label1 = QLabel('Количество файлов:', self)
        self.label1.move(10, 10)
        self.lineEdit1 = QLineEdit(self)
        self.lineEdit1.move(10, 30)

        self.label2 = QLabel('Сколько необходимо сформировать данных в файле:', self)
        self.label2.move(10, 60)
        self.lineEdit2 = QLineEdit(self)
        self.lineEdit2.move(10, 80)

        self.progress = QProgressBar(self)
        self.progress.setGeometry(10, 150, 350, 25)

        self.button = QPushButton('Сформировать', self)
        self.button.move(10, 120)

        self.button.clicked.connect(lambda: self.generate_excel_files())

    def generate_excel_files(self):
        try:
            n_people = self.lineEdit1.text()
            n_data = self.lineEdit2.text()
            # Проверяем, что введены числа
            if not n_people.isdigit() or not n_data.isdigit():
                QMessageBox.warning(self, 'Ошибка',
                                    'В полях должны быть указаны целые числа')
                return
            n_people = int(n_people)
            n_data = int(n_data)
        except ValueError:
            return

        self.progress.setMaximum(n_people)

        for i in range(n_people):
            filename = f'data_{i+1}.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Данные'

            headers = ['снилс', 'ИНН']
            for col, header in enumerate(headers):
                worksheet.cell(row=1, column=col+1, value=header)

            for row in range(2, n_data+2):
                snils = get_snils()
                inn = get_random_inn_fl()
                worksheet.cell(row=row, column=1, value=snils)
                worksheet.cell(row=row, column=2, value=inn)

            if not os.path.exists('data'):
                os.makedirs('data')

            workbook.save(f'data/{filename}')

            self.progress.setValue(i + 1)

        QMessageBox.information(self, 'Готово', 'Excel-файлы успешно сформированы')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    excel_form = ExcelForm()
    excel_form.show()
    sys.exit(app.exec())
